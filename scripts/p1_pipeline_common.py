from __future__ import annotations

import csv
import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "p1"
OUTPUT_DIR = ROOT / "outputs" / "p1"
WORKBOOK_PATH = ROOT / "Lattice crashworthiness data.xlsx"
STRUCTURE_MAPPING_PATH = ROOT / "mappings" / "structure" / "structure_mapping.csv"
MATERIAL_MAPPING_PATH = ROOT / "mappings" / "material" / "material_mapping.csv"
PROCESS_MAPPING_PATH = ROOT / "mappings" / "process" / "process_mapping.csv"
DEFAULT_WORKBOOK_ROW = 104

SEED_INPUT_PATH = DATA_DIR / "seed_input.json"
SEED_SNAPSHOT_PATH = OUTPUT_DIR / "seed" / "seed_input.json"
PARSED_PATH = OUTPUT_DIR / "parsed" / "parsed_sample.json"
SAMPLES_PATH = OUTPUT_DIR / "extracted" / "samples_v1.json"
EVIDENCE_PATH = OUTPUT_DIR / "extracted" / "evidence_v1.json"
REPORT_PATH = OUTPUT_DIR / "review" / "p1_chain_report.json"

SAMPLE_SCHEMA_PATH = ROOT / "schemas" / "samples" / "schema_v1.json"
EVIDENCE_SCHEMA_PATH = ROOT / "schemas" / "evidence" / "schema_v1.json"


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    ensure_parent(path)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "–", "—"} or text.lower() in {"n/a", "na", "null"}:
            return None
        return text
    text = str(value).strip()
    return text or None


def _normalize_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _normalize_text(value)
    if text is None:
        return None
    text = text.replace("≈", "").replace("~", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_string_or_none(value: Any) -> str | None:
    text = _normalize_text(value)
    return text if text is not None else None


def _slugify(value: str) -> str:
    text = re.sub(r"[^0-9a-zA-Z]+", "-", value.strip().lower())
    return re.sub(r"-+", "-", text).strip("-")


def _normalize_lookup_key(value: Any) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    return re.sub(r"\s+", " ", text).lower()


def normalize_doi(raw_doi: Any) -> str | None:
    text = _normalize_text(raw_doi)
    if text is None:
        return None
    lowered = text.lower()
    prefixes = (
        "https://doi.org/",
        "http://doi.org/",
        "doi.org/",
        "doi:",
    )
    for prefix in prefixes:
        if lowered.startswith(prefix):
            text = text[len(prefix) :]
            break
    return text.strip() or None


@lru_cache(maxsize=1)
def _load_csv_rows(path: str) -> tuple[dict[str, str], ...]:
    rows: list[dict[str, str]] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append({key: (value.strip() if isinstance(value, str) else value) for key, value in row.items()})
    return tuple(rows)


@lru_cache(maxsize=1)
def _material_mapping() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for row in _load_csv_rows(str(MATERIAL_MAPPING_PATH)):
        key = _normalize_lookup_key(row.get("raw_material"))
        if key:
            lookup[key] = row
    return lookup


@lru_cache(maxsize=1)
def _process_mapping() -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for row in _load_csv_rows(str(PROCESS_MAPPING_PATH)):
        key = _normalize_lookup_key(row.get("raw_process"))
        if key:
            lookup[key] = row
    return lookup


def normalize_material_canonical(raw_material: Any) -> str | None:
    key = _normalize_lookup_key(raw_material)
    if key is None:
        return None
    row = _material_mapping().get(key)
    if row:
        return _normalize_string_or_none(row.get("material_canonical"))
    text = _normalize_text(raw_material)
    if text is None:
        return None
    if text.lower() in {"aluminium", "aluminum"}:
        return "Aluminum"
    if text.lower() in {"316l", "304l", "304 ss"}:
        return text.upper() if text.lower().endswith("l") else text
    return text


def normalize_material_family(raw_material: Any) -> str:
    key = _normalize_lookup_key(raw_material)
    if key is not None:
        row = _material_mapping().get(key)
        if row:
            family = _normalize_string_or_none(row.get("material_family"))
            if family:
                return family
    text = _normalize_text(raw_material)
    if text is None:
        return "unknown"
    normalized = text.lower()
    if normalized in {"316l", "304l", "304 ss", "17-4 ph stainless steel"}:
        return "stainless_steel"
    if normalized in {"aluminium", "aluminum"}:
        return "aluminum_alloy"
    if normalized in {"ti-6al-4v"}:
        return "titanium_alloy"
    if normalized in {"nylon 12", "nylon-12", "pa12", "pa-12", "polyamide 12"}:
        return "polymer"
    return "unknown"


def normalize_process_canonical(raw_process: Any) -> str | None:
    key = _normalize_lookup_key(raw_process)
    if key is None:
        return None
    row = _process_mapping().get(key)
    if row:
        return _normalize_string_or_none(row.get("process_canonical"))
    text = _normalize_text(raw_process)
    return text


def normalize_process_family(raw_process: Any) -> str:
    key = _normalize_lookup_key(raw_process)
    if key is not None:
        row = _process_mapping().get(key)
        if row:
            family = _normalize_string_or_none(row.get("process_family"))
            if family:
                return family
    text = _normalize_text(raw_process)
    if text is None:
        return "unknown"
    normalized = text.lower()
    if normalized == "selective laser melting":
        return "slm"
    mapping = {
        "slm": "slm",
        "sls": "sls",
        "sla": "sla",
        "fdm": "fdm",
        "forming and assembly": "forming_and_assembly",
    }
    return mapping.get(normalized, "unknown")


def normalize_structure_main_class(raw_design: Any, raw_structure: Any, raw_type: Any | None = None) -> str:
    normalized_type = _normalize_lookup_key(raw_type)
    type_map = {
        "tpms": "tpms",
        "general 2d": "honeycomb_2d",
        "2d general": "honeycomb_2d",
        "truss": "truss_lattice",
        "plate": "plate_lattice",
        "tubular": "tube_lattice",
        "bioinspired": "bioinspired",
        "voronoi": "voronoi",
        "hybrid": "hybrid_lattice",
        "unknown": "unknown",
    }
    if normalized_type in type_map:
        return type_map[normalized_type]

    for candidate in (raw_design, raw_structure):
        normalized = _normalize_lookup_key(candidate)
        if normalized is None:
            continue
        if normalized in {"gyroid", "diamond"}:
            return "tpms"
        if normalized in {"bcc", "bccz", "octet"}:
            return "truss_lattice"
        if normalized in {"hexagonal", "re-entrant", "reentrant", "sine-wave", "sine wave", "beam", "kagome"}:
            return "honeycomb_2d"
        if normalized == "primitive":
            return "tpms"
        if normalized == "bioinspired":
            return "bioinspired"
        if normalized == "voronoi":
            return "voronoi"
        if normalized == "plate":
            return "plate_lattice"
        if normalized == "tube":
            return "tube_lattice"
    return "unknown"


def normalize_structure_subtype_list(raw_design: Any, raw_structure: Any, raw_type: Any | None = None) -> list[str]:
    for candidate in (raw_design, raw_structure, raw_type):
        normalized = _normalize_lookup_key(candidate)
        if normalized is None:
            continue
        if normalized == "general 2d":
            return ["general_2d"]
        if normalized == "2d general":
            return ["general_2d"]
        return [_slugify(normalized)]
    return []


def normalize_analysis_type(raw_analysis: Any) -> str:
    text = _normalize_lookup_key(raw_analysis)
    if text is None:
        return "unknown"
    if text in {"experimental", "exp"}:
        return "experimental"
    if text in {"numerical", "simulation"}:
        return "numerical"
    if text in {"experimental/numerical", "experimental numerical", "experimental_numerical"}:
        return "experimental_numerical"
    return "unknown"


def normalize_test_mode(raw_test_mode: Any) -> str:
    text = _normalize_lookup_key(raw_test_mode)
    if text is None:
        return "unknown"
    if text in {"quasi-static", "quasi static", "quasi_static"}:
        return "quasi_static"
    if text == "impact":
        return "impact"
    if text == "dynamic":
        return "dynamic"
    return "unknown"


def normalize_loading_direction(raw_direction: Any) -> str:
    text = _normalize_lookup_key(raw_direction)
    if text is None:
        return "unknown"
    if text == "uniaxial":
        return "uniaxial"
    if text == "biaxial":
        return "biaxial"
    if text == "lateral":
        return "lateral"
    if text in {"concentrated load", "concentraded load"}:
        return "concentrated_load"
    return "unknown"


def _read_database_row(workbook_path: Path, row_number: int) -> dict[int, Any]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook["Database"]
        return {column: worksheet.cell(row_number, column).value for column in range(1, 28)}
    finally:
        workbook.close()


def _count_rows_with_doi(workbook_path: Path, normalized_doi: str) -> int:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook["Database"]
        count = 0
        for row_number in range(3, worksheet.max_row + 1):
            candidate = normalize_doi(worksheet.cell(row_number, 3).value)
            if candidate == normalized_doi:
                count += 1
        return count
    finally:
        workbook.close()


def build_seed_input_from_workbook(workbook_path: Path | str = WORKBOOK_PATH, row_number: int = DEFAULT_WORKBOOK_ROW) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    row = _read_database_row(workbook_path, row_number)

    doi = normalize_doi(row[3])
    if doi is None:
        raise ValueError(f"Row {row_number} does not contain a DOI")

    paper_id = _slugify(doi)
    raw_design = _normalize_string_or_none(row[4])
    raw_structure = _normalize_string_or_none(row[5])
    raw_type = _normalize_string_or_none(row[6])
    structure_main_class = normalize_structure_main_class(raw_design, raw_structure, raw_type)
    structure_subtype_list = normalize_structure_subtype_list(raw_design, raw_structure, raw_type)
    normalized_structure = "-".join([structure_main_class, *structure_subtype_list]) if structure_subtype_list else structure_main_class

    sample_label_in_paper = f"row-{row_number}"
    is_multi_sample_paper = _count_rows_with_doi(workbook_path, doi) > 1

    return {
        "paper_id": paper_id,
        "sample_id": f"{paper_id}-{sample_label_in_paper}-{normalized_structure}-{row_number}",
        "doi": doi,
        "title": _normalize_string_or_none(row[2]),
        "sample_label_in_paper": sample_label_in_paper,
        "is_multi_sample_paper": is_multi_sample_paper,
        "raw_design": raw_design,
        "raw_structure": raw_structure,
        "raw_type": raw_type,
        "raw_material": _normalize_string_or_none(row[9]),
        "raw_material_group": _normalize_string_or_none(row[10]),
        "raw_process": _normalize_string_or_none(row[13]),
        "analysis_type": normalize_analysis_type(row[14]),
        "test_mode": normalize_test_mode(row[15]),
        "loading_direction": normalize_loading_direction(row[16]),
        "velocity_m_s_raw": _normalize_string_or_none(row[17]),
        "velocity_m_s": _normalize_float(row[17]),
        "sea_j_g_raw": _normalize_string_or_none(row[20]),
        "sea_j_g": _normalize_float(row[20]),
        "pcf_kn_raw": _normalize_string_or_none(row[22]),
        "pcf_kn": _normalize_float(row[22]),
        "mcf_kn_raw": _normalize_string_or_none(row[24]),
        "mcf_kn": _normalize_float(row[24]),
        "cfe_raw": _normalize_string_or_none(row[26]),
        "cfe": _normalize_float(row[26]),
        "plateau_stress_mpa_raw": _normalize_string_or_none(row[27]),
        "plateau_stress_mpa": _normalize_float(row[27]),
        "relative_density_raw": _normalize_string_or_none(row[8]),
        "relative_density_value": _normalize_float(row[8]),
    }


def build_seed_input(workbook_path: Path | str = WORKBOOK_PATH, row_number: int = DEFAULT_WORKBOOK_ROW) -> dict[str, Any]:
    return build_seed_input_from_workbook(workbook_path, row_number=row_number)


def build_parsed_artifact(seed: dict[str, Any]) -> dict[str, Any]:
    return {
        "pipeline": "p1-first-minimal-chain",
        "paper_id": seed["paper_id"],
        "sample_id": seed["sample_id"],
        "raw_input": seed,
        "normalized_fields": {
            "structure_main_class": normalize_structure_main_class(seed["raw_design"], seed["raw_structure"], seed["raw_type"]),
            "structure_subtype_list": normalize_structure_subtype_list(seed["raw_design"], seed["raw_structure"], seed["raw_type"]),
            "material_canonical": normalize_material_canonical(seed["raw_material"]),
            "material_family": normalize_material_family(seed["raw_material"]),
            "process_canonical": normalize_process_canonical(seed["raw_process"]),
            "process_family": normalize_process_family(seed["raw_process"]),
            "analysis_type": seed["analysis_type"],
            "test_mode": seed["test_mode"],
            "loading_direction": seed["loading_direction"],
        },
        "alignment": [
            {
                "field_name": "raw_design",
                "resolved_to": "structure_main_class",
                "source_priority": "T1",
                "source_type": "text",
            },
            {
                "field_name": "raw_material",
                "resolved_to": "material_family",
                "source_priority": "T1",
                "source_type": "text",
            },
            {
                "field_name": "raw_process",
                "resolved_to": "process_family",
                "source_priority": "T1",
                "source_type": "text",
            },
            {
                "field_name": "sea_j_g",
                "resolved_to": "sea_j_g",
                "source_priority": "T2",
                "source_type": "derived",
            },
        ],
    }


def build_sample_record(seed: dict[str, Any]) -> dict[str, Any]:
    return {
        "paper_id": seed["paper_id"],
        "sample_id": seed["sample_id"],
        "doi": seed["doi"],
        "title": seed["title"],
        "sample_label_in_paper": seed["sample_label_in_paper"],
        "is_multi_sample_paper": seed["is_multi_sample_paper"],
        "raw_design": seed["raw_design"],
        "raw_structure": seed["raw_structure"],
        "raw_type": seed["raw_type"],
        "structure_main_class": normalize_structure_main_class(seed["raw_design"], seed["raw_structure"], seed["raw_type"]),
        "structure_subtype_list": normalize_structure_subtype_list(seed["raw_design"], seed["raw_structure"], seed["raw_type"]),
        "is_hierarchical": False,
        "is_graded": False,
        "is_optimized": False,
        "relative_density_raw": seed["relative_density_raw"],
        "relative_density_value": seed["relative_density_value"],
        "raw_material": seed["raw_material"],
        "raw_material_group": seed["raw_material_group"],
        "material_canonical": normalize_material_canonical(seed["raw_material"]),
        "material_family": normalize_material_family(seed["raw_material"]),
        "raw_process": seed["raw_process"],
        "process_canonical": normalize_process_canonical(seed["raw_process"]),
        "process_family": normalize_process_family(seed["raw_process"]),
        "analysis_type": seed["analysis_type"],
        "test_mode": seed["test_mode"],
        "loading_direction": seed["loading_direction"],
        "velocity_m_s_raw": seed["velocity_m_s_raw"],
        "velocity_m_s": seed["velocity_m_s"],
        "sea_j_g_raw": seed["sea_j_g_raw"],
        "sea_j_g": seed["sea_j_g"],
        "pcf_kn_raw": seed["pcf_kn_raw"],
        "pcf_kn": seed["pcf_kn"],
        "mcf_kn_raw": seed["mcf_kn_raw"],
        "mcf_kn": seed["mcf_kn"],
        "cfe_raw": seed["cfe_raw"],
        "cfe": seed["cfe"],
        "plateau_stress_mpa_raw": seed["plateau_stress_mpa_raw"],
        "plateau_stress_mpa": seed["plateau_stress_mpa"],
        "confidence_overall": 0.88,
        "needs_manual_review": True,
        "review_status": "pending",
        "review_notes": "Workbook-derived seed used to validate the first minimal chain.",
    }


def build_evidence_records(seed: dict[str, Any], sample: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "evidence_id": f"{seed['sample_id']}-ev-001",
            "sample_id": sample["sample_id"],
            "field_name": "raw_design",
            "field_value": seed["raw_design"],
            "source_priority": "T1",
            "source_type": "table",
            "page_no": None,
            "figure_id": None,
            "table_id": "Database",
            "text_snippet": f"Database row records raw_design={seed['raw_design']}.",
            "extractor": "p1_workbook_adapter",
            "extract_confidence": 1.0,
            "verified_by_human": False,
        },
        {
            "evidence_id": f"{seed['sample_id']}-ev-002",
            "sample_id": sample["sample_id"],
            "field_name": "raw_material",
            "field_value": seed["raw_material"],
            "source_priority": "T1",
            "source_type": "table",
            "page_no": None,
            "figure_id": None,
            "table_id": "Database",
            "text_snippet": f"Database row records raw_material={seed['raw_material']}.",
            "extractor": "p1_workbook_adapter",
            "extract_confidence": 1.0,
            "verified_by_human": False,
        },
        {
            "evidence_id": f"{seed['sample_id']}-ev-003",
            "sample_id": sample["sample_id"],
            "field_name": "raw_process",
            "field_value": seed["raw_process"],
            "source_priority": "T1",
            "source_type": "table",
            "page_no": None,
            "figure_id": None,
            "table_id": "Database",
            "text_snippet": f"Database row records raw_process={seed['raw_process']}.",
            "extractor": "p1_workbook_adapter",
            "extract_confidence": 1.0,
            "verified_by_human": False,
        },
        {
            "evidence_id": f"{seed['sample_id']}-ev-004",
            "sample_id": sample["sample_id"],
            "field_name": "sea_j_g",
            "field_value": str(seed["sea_j_g"]),
            "source_priority": "T1",
            "source_type": "table",
            "page_no": None,
            "figure_id": None,
            "table_id": "Database",
            "text_snippet": "Database row records SEA(J/g).",
            "extractor": "p1_workbook_adapter",
            "extract_confidence": 1.0,
            "verified_by_human": False,
        },
    ]


def validate_against_schema(instance: Any, schema: dict[str, Any], path: str = "$") -> list[str]:
    errors: list[str] = []

    def fail(message: str) -> None:
        errors.append(f"{path}: {message}")

    def check(value: Any, rule: dict[str, Any], current_path: str) -> None:
        allowed_types = rule.get("type")
        if isinstance(allowed_types, str):
            allowed_types = [allowed_types]
        if allowed_types is not None:
            if value is None:
                if "null" not in allowed_types:
                    errors.append(f"{current_path}: expected {allowed_types}, got null")
                    return
            else:
                if "string" in allowed_types and isinstance(value, str):
                    pass
                elif "boolean" in allowed_types and isinstance(value, bool):
                    pass
                elif "integer" in allowed_types and isinstance(value, int) and not isinstance(value, bool):
                    pass
                elif "number" in allowed_types and isinstance(value, (int, float)) and not isinstance(value, bool):
                    pass
                elif "array" in allowed_types and isinstance(value, list):
                    pass
                elif "object" in allowed_types and isinstance(value, dict):
                    pass
                else:
                    errors.append(f"{current_path}: type mismatch, expected {allowed_types}, got {type(value).__name__}")
                    return

        if value is None:
            return

        if "enum" in rule and value not in rule["enum"]:
            errors.append(f"{current_path}: value {value!r} not in enum {rule['enum']}")
            return

        if isinstance(value, str):
            min_length = rule.get("minLength")
            if min_length is not None and len(value) < min_length:
                errors.append(f"{current_path}: string shorter than minLength {min_length}")
            max_length = rule.get("maxLength")
            if max_length is not None and len(value) > max_length:
                errors.append(f"{current_path}: string longer than maxLength {max_length}")
            return

        if isinstance(value, bool):
            return

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            minimum = rule.get("minimum")
            if minimum is not None and value < minimum:
                errors.append(f"{current_path}: value {value} below minimum {minimum}")
            maximum = rule.get("maximum")
            if maximum is not None and value > maximum:
                errors.append(f"{current_path}: value {value} above maximum {maximum}")
            return

        if isinstance(value, list):
            items_rule = rule.get("items")
            if items_rule:
                for index, item in enumerate(value):
                    check(item, items_rule, f"{current_path}[{index}]")
            return

        if isinstance(value, dict):
            properties = rule.get("properties", {})
            required = rule.get("required", [])
            for key in required:
                if key not in value:
                    errors.append(f"{current_path}: missing required property {key!r}")
            if rule.get("additionalProperties") is False:
                for key in value:
                    if key not in properties:
                        errors.append(f"{current_path}: unexpected property {key!r}")
            for key, subrule in properties.items():
                if key in value:
                    check(value[key], subrule, f"{current_path}.{key}")
            return

    check(instance, schema, path)
    return errors


def validate_instance_file(instance_path: Path, schema_path: Path, label: str) -> list[str]:
    instance = read_json(instance_path)
    schema = read_json(schema_path)
    if isinstance(instance, list):
        if not instance:
            return [f"{label}: empty collection"]
        errors: list[str] = []
        for index, item in enumerate(instance):
            item_errors = validate_against_schema(item, schema, path=f"{label}[{index}]")
            errors.extend(item_errors)
        return errors
    errors = validate_against_schema(instance, schema, path=label)
    return errors
